"""
excel_reader.py - Lectura y normalización del archivo Excel (.xlsx).

Maneja:
- Lectura de metadatos institucionales (siglo y número de acervo) desde las filas 4 y 7.
- Encabezados institucionales (skiprows)
- Columnas con sub-encabezados (FECHA INICIAL / FECHA FINAL dentro de DATA CRÓNICA)
- Filas completamente vacías
- Filas de marcador de sección ('Protocolo N°X', 'Registro N°X') intercaladas en los datos
"""
import logging
import re
from pathlib import Path
from typing import Optional
import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)


def _normalize_col(name: str) -> str:
    """Normaliza un nombre de columna: strip de espacios extremos (conserva saltos de línea internos)."""
    return str(name).strip()


def load_excel_metadata(
    excel_path: Path,
    meta_row_siglo: int = 4,
    meta_row_acervo: int = 7,
) -> dict:
    """
    Lee los metadatos globales del fondo desde las filas de cabecera del Excel.

    Estructura esperada:
      - Fila 4: «Seccion: XVI»          → extrae 'XVI'
      - Fila 7: «Código del fondo: N7»  → extrae '7'

    Args:
        excel_path:      Ruta al archivo .xlsx
        meta_row_siglo:  Número de fila real (1-indexed) con el siglo.
        meta_row_acervo: Número de fila real (1-indexed) con el código de acervo.

    Returns:
        Dict con claves:
          'siglo'      : str, ej. 'XVI'
          'acervo_num' : str, ej. '7'
          'raw_siglo'  : str, contenido original de la celda
          'raw_acervo' : str, contenido original de la celda
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Archivo Excel no encontrado: {path}")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    raw_siglo  = str(ws.cell(meta_row_siglo,  1).value or "").strip()
    raw_acervo = str(ws.cell(meta_row_acervo, 1).value or "").strip()
    wb.close()

    # Extraer siglo: «Seccion: XVI» → 'XVI'
    m_siglo = re.search(r':\s*(.+)$', raw_siglo)
    siglo = m_siglo.group(1).strip() if m_siglo else raw_siglo or "Sin_Siglo"

    # Extraer número de acervo: «Código del fondo: N7» → '7'
    m_num = re.search(r'N(\d+)', raw_acervo, re.IGNORECASE)
    if m_num:
        acervo_num = m_num.group(1)
    else:
        # Fallback: todo lo que sigue al ':'
        m_colon = re.search(r':\s*(.+)$', raw_acervo)
        acervo_num = m_colon.group(1).strip() if m_colon else raw_acervo or "Sin_Num"

    logger.info(f"Metadata leida -> Siglo: {siglo!r}  |  Acervo N.: {acervo_num!r}")
    return {
        "siglo":      siglo,
        "acervo_num": acervo_num,
        "raw_siglo":  raw_siglo,
        "raw_acervo": raw_acervo,
    }


def load_excel(
    excel_path: Path,
    skip_rows: int = 7,
    col_fecha_ini: str = "FECHA INICIAL",
) -> pd.DataFrame:
    """
    Carga el Excel y devuelve un DataFrame limpio y listo para procesar.

    El índice del DataFrame corresponde al número de fila REAL del Excel
    (1-indexed, tal como lo ve el usuario en la hoja de cálculo).

    Estructura del Excel:
      - Filas 1..skip_rows: encabezados institucionales (se omiten)
      - Fila skip_rows+1 (fila 8): encabezado real de columnas
      - Fila skip_rows+2 (fila 9): sub-encabezados FECHA INICIAL / FECHA FINAL
      - Filas skip_rows+3 en adelante: datos reales + marcadores de sección

    Marcadores de sección intercalados en los datos
    (ej. «Protocolo N°16», «Registro N°1») se eliminan automáticamente:
    son filas cuyo valor en la primera columna (N° DE REGISTRO) no es numérico.

    Args:
        excel_path:    Ruta al archivo .xlsx
        skip_rows:     Número de filas a saltar (0-indexed) antes del header.
        col_fecha_ini: Nombre a asignar a la columna de fecha inicial.

    Returns:
        DataFrame con columnas normalizadas, filas vacías y marcadores eliminados,
        y el índice = número de fila real en el Excel (1-indexed).
    """
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Archivo Excel no encontrado: {path}")

    logger.info(f"Cargando Excel: {path}")

    df = pd.read_excel(
        path,
        skiprows=skip_rows,
        header=0,
        engine="openpyxl",
        dtype=str,           # Todo como string; evita conversiones numéricas inesperadas
    )

    # Normalizar nombres de columna (strip de espacios extremos)
    df.columns = [_normalize_col(c) for c in df.columns]

    # Eliminar columnas completamente sin nombre (Unnamed: X)
    unnamed = [c for c in df.columns if str(c).startswith("Unnamed:")]
    if unnamed:
        df.drop(columns=unnamed, inplace=True)

    # El Excel tiene dos sub-columnas bajo "DATA CRÓNICA".
    # Después del drop de Unnamed, DATA CRÓNICA puede aparecer duplicada como "DATA CRÓNICA.1".
    rename_map = {}
    # Buscar columnas que contengan DATA CRÓNICA (con o sin tilde, con o sin sufijo)
    data_cronica_cols = [c for c in df.columns if "DATA CR" in c.upper()]
    if len(data_cronica_cols) >= 2:
        rename_map[data_cronica_cols[0]] = col_fecha_ini
        rename_map[data_cronica_cols[1]] = "FECHA FINAL"
    elif len(data_cronica_cols) == 1:
        rename_map[data_cronica_cols[0]] = col_fecha_ini
    df.rename(columns=rename_map, inplace=True)

    # Asignar el número de fila REAL del Excel como índice.
    # skip_rows filas omitidas + 1 fila de header = skip_rows+1 antes de datos.
    # La primera fila de datos en pandas (índice 0) = fila skip_rows+2 del Excel.
    df.index = df.index + skip_rows + 2  # índice = fila real del Excel

    # Eliminar filas completamente vacías (conservando el índice real)
    df.dropna(how="all", inplace=True)

    # Reemplazar valores 'nan' string por vacío
    df = df.fillna("")

    # ── Eliminar filas de marcador de sección ──────────────────────────────────
    # Son filas como «Protocolo N°16» o «Registro N°1» en la columna N° DE REGISTRO.
    # Criterio: la primera columna de datos NO es numérica (no es un número de registro).
    first_col = df.columns[0]
    _section_marker_re = re.compile(
        r'^\s*(protocolo|registro)\b', re.IGNORECASE
    )

    def _is_data_row(val: str) -> bool:
        v = str(val).strip()
        if not v:
            return False
        # Filas de datos reales tienen un número (o texto breve) como n° de registro
        if _section_marker_re.match(v):
            return False
        return True

    before = len(df)
    df = df[df[first_col].apply(_is_data_row)]
    removed = before - len(df)
    if removed:
        logger.info(f"Eliminados {removed} marcadores de sección (Protocolo/Registro) de los datos")

    logger.info(f"Excel cargado: {len(df)} registros (filas {df.index.min()}–{df.index.max()} del Excel)")
    return df
